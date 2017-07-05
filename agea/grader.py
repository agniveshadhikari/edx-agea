
from openpyxl import *
import logging

log = logging.getLogger(__name__)



def grade(questionPath, answerPath, solutionPath):
    annotated = Workbook()
    sample = Workbook()
    answer = Workbook()
    log.info("-------------")
    log.info(questionPath)
    answer=load_workbook(answerPath)
    numby=0
    for sheet in answer.worksheets:
        numby+=1
    if numby != 1:
        return -1

    annotated = load_workbook(questionPath).active
    sample = load_workbook(solutionPath).active
    answer = load_workbook(answerPath).active

    score = 0
    
    for row in annotated.iter_rows():
        for cell in row:
            coord = cell.coordinate
            if str(annotated[coord].value)[0] == '<':
                if answer[coord].value == sample[coord].value:
                    score += int((annotated[coord].value)[1:-1])

    return score
def total_marks(questionPath):
    question = Workbook()
    question = load_workbook(questionPath).active
    score = 0
    for row in question.iter_rows():
        for cell in row:
            coord = cell.coordinate
            if str(question[coord].value)[0] == '<':
                score += int((question[coord].value)[1:-1])

    return score

"""
def grade(questionPath, answerPath, solutionPath):
    return 69
"""

